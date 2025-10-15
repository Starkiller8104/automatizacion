python
import io
import datetime as dt
import pytz
import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

APP_TITLE = "TC USD/MXN (FIX) • Paso 1"
MX_TZ = pytz.timezone("America/Mexico_City")
SIE_URL = "https://www.banxico.org.mx/SieAPIRest/service/v1/series/{series}/datos/{start}/{end}"

def today_mx() -> dt.date:
    return dt.datetime.now(MX_TZ).date()

def date_to_iso(d: dt.date) -> str:
    return d.strftime("%Y-%m-%d")

def last_value_leq_date(series_id: str, token: str, target_date: dt.date, lookback_days: int = 10):
    """
    Devuelve (fecha_valor, valor) del último dato disponible <= target_date.
    Maneja fines de semana/feriados.
    """
    start = target_date - dt.timedelta(days=lookback_days)
    url = SIE_URL.format(series=series_id, start=date_to_iso(start), end=date_to_iso(target_date))
    headers = {"Bmx-Token": token.strip()}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    js = r.json()
    series = js.get("bmx", {}).get("series", [])
    datos = series[0].get("datos", []) if series else []
    if not datos:
        return None, None
    df = pd.DataFrame(datos)
    if df.empty:
        return None, None
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce").dt.date
    df["dato"] = pd.to_numeric(df["dato"].replace({"N/E": None}), errors="coerce")
    df = df.dropna(subset=["dato"]).sort_values("fecha")
    if df.empty:
        return None, None
    row = df.tail(1).iloc[0]
    return row["fecha"], float(row["dato"])

def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="💵", layout="centered")
    st.title(APP_TITLE)
    st.caption("Objetivo: validar la lógica para obtener USD/MXN FIX (Banxico) para 'hoy' y 'ayer'.")

    # Parámetros
    default_token = st.secrets.get("BANXICO_TOKEN", "") if hasattr(st, "secrets") else ""
    if not default_token:
        default_token = "677aaedf11d11712aa2ccf73da4d77b6b785474eaeb2e092f6bad31b29de6609"  # puedes borrar si usarás secrets
    token = st.text_input("BANXICO_TOKEN", value=default_token, type="password")

    colA, colB = st.columns(2)
    hoy = colA.date_input("Fecha D (hoy)", value=today_mx())
    ayer = colB.date_input("Fecha C (ayer)", value=hoy - dt.timedelta(days=1))

    series_fix = st.text_input("Serie SIE USD/MXN FIX", value="SF43718",
                               help="Clave SIE para FIX (ejemplo común: SF43718).")

    if st.button("Consultar USD FIX"):
        if not token.strip():
            st.error("Ingresa el BANXICO_TOKEN.")
            st.stop()

        with st.spinner("Consultando Banxico..."):
            fecha_c, fix_c = last_value_leq_date(series_fix, token, ayer)
            fecha_d, fix_d = last_value_leq_date(series_fix, token, hoy)

        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Columna C (ayer)")
            st.write(f"Fecha solicitada: {date_to_iso(ayer)}")
            st.write(f"Fecha con dato: {date_to_iso(fecha_c) if fecha_c else '—'}")
            st.metric("USD/MXN FIX (C)", f"{fix_c:.4f}" if fix_c is not None else "N/D")
        with c2:
            st.subheader("Columna D (hoy)")
            st.write(f"Fecha solicitada: {date_to_iso(hoy)}")
            st.write(f"Fecha con dato: {date_to_iso(fecha_d) if fecha_d else '—'}")
            st.metric("USD/MXN FIX (D)", f"{fix_d:.4f}" if fix_d is not None else "N/D")

        # --- Escribir directamente en la plantilla (opcional en este paso) ---
        st.divider()
        st.subheader("Escribir DIRECTO en tu plantilla (opcional en este paso)")
        st.caption("Se escribirá: C2=ayer, D2=hoy, C5=FIX(ayer), D5=FIX(hoy)")
        tpl = st.file_uploader("Sube tu plantilla .xlsx", type=["xlsx"])

        if tpl and (fix_c is not None or fix_d is not None):
            if st.button("Escribir en plantilla y descargar"):
                try:
                    raw = tpl.read()
                    wb = load_workbook(io.BytesIO(raw))
                    ws = wb.active  # primera hoja

                    # Fechas
                    ws["C2"] = date_to_iso(ayer)
                    ws["D2"] = date_to_iso(hoy)

                    # Valores FIX
                    ws["C5"] = float(fix_c) if fix_c is not None else None
                    ws["D5"] = float(fix_d) if fix_d is not None else None

                    out = io.BytesIO()
                    wb.save(out)
                    st.success("Plantilla actualizada correctamente.")
                    st.download_button(
                        "Descargar plantilla actualizada",
                        data=out.getvalue(),
                        file_name=f"Plantilla_TC_USD_{date_to_iso(hoy)}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Error al escribir en la plantilla: {e}")

    st.caption("Si valida bien, en el siguiente paso integramos más divisas/series y celdas específicas.")

if __name__ == "__main__":
    main()
