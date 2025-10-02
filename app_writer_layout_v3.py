
# app_writer_layout_v3.py
from openpyxl import load_workbook
from datetime import datetime, timedelta
from typing import Dict, Any

NO_MODIFY = set("B2 B5 B10 B14 B18 B21 B22 B23 B24 B26 B27 B28 B29 B30".split())
BLOCKS = {
    "DOLAR": list(range(5, 8)),   # 5: USD/MXN; 6: Compra; 7: Venta
    "YEN": list(range(10, 12)),
    "EURO": list(range(14, 16)),
    "UDIS": [18],
    "TIIE": list(range(21, 25)),
    "CETES": list(range(27, 31)),
    "UMA": list(range(33, 36)),
}
DATE_ROW = 2
DATE_COLS = ["B","C","D","E","F","G"]  # 6 días hábiles; G es hoy

def _has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.lstrip().startswith("=")

def _safe_set(ws, addr: str, value: Any):
    if addr in NO_MODIFY:
        return
    c = ws[addr]
    if _has_formula(c):
        return
    ws[addr].value = value

def set_dates_b2g(ws, dates):
    # 'dates' debe tener longitud 6, con la última como "hoy"
    for col, d in zip(DATE_COLS, dates):
        c = ws[f"{col}{DATE_ROW}"]
        if not _has_formula(c):
            ws[f"{col}{DATE_ROW}"].value = d

def enforce_formats(ws):
    # Porcentajes en G21:G24 y G27:G30
    for r in (*range(21,25), *range(27,31)):
        cell = ws[f"G{r}"]
        v = cell.value
        if isinstance(v, (int,float)) and v is not None and v >= 1.0:
            cell.value = v/100.0
        cell.number_format = "0.00%"
    # 4 decimales
    for addr in ["G15","G18"]:
        ws[addr].number_format = "0.0000"

def fill_payload(ws, payload: Dict[str, Dict[str, Dict[str, Any]]]):
    for block, rows in (payload or {}).items():
        valid = set(BLOCKS.get(block, []))
        for r_str, cols in rows.items():
            try:
                r = int(r_str)
            except:
                continue
            if r not in valid:
                continue
            for col, val in cols.items():
                if col not in ("F","G"):
                    continue
                _safe_set(ws, f"{col}{r}", val)

def write_layout_v3(template_path: str, out_path: str, header_dates=None, payload=None):
    wb = load_workbook(template_path, data_only=False, keep_vba=False)
    ws = wb["Indicadores"]
    if header_dates:
        set_dates_b2g(ws, header_dates)
    if payload:
        fill_payload(ws, payload)
    enforce_formats(ws)
    wb.save(out_path)
    return out_path
